"""
Spreadsheet import parser — CSV and Excel (.xlsx) → assignment dicts.

Pure parsing/normalization layer. No database I/O. Tolerates messy real-world
files via flexible column aliases and per-row error collection.
"""

from __future__ import annotations

import csv
import hashlib
import io
import re
from typing import Any, Dict, List, Optional, Tuple

from config.logging_config import get_logger

logger = get_logger(__name__)

# Flexible header aliases (lowercase, stripped). First match wins per field.
FIELD_ALIASES: Dict[str, List[str]] = {
    "assignment_id": [
        "assignment_id", "id", "project_id", "project id", "projectid",
        "key", "code", "slug",
    ],
    "name": [
        "name", "assignment_name", "assignment name", "project_name",
        "project name", "project", "title", "engagement",
    ],
    "description": ["description", "desc", "summary", "notes", "details"],
    "team_size": ["team_size", "team size", "team", "headcount", "size", "fte"],
    "monthly_burn_rate": [
        "monthly_burn_rate", "monthly burn", "monthly_burn", "burn",
        "burn_rate", "spend", "monthly_spend", "budget", "monthly budget",
    ],
    "target_monthly_burn": [
        "target_monthly_burn", "target burn", "target_burn", "target_budget",
        "target budget", "budget_target", "target monthly burn",
    ],
    "status": ["status", "state", "phase"],
}

CONNECTOR_COLUMNS = {
    "github": ["github", "enable_github", "github_enabled"],
    "jira": ["jira", "enable_jira", "jira_enabled"],
    "aws": ["aws", "enable_aws", "aws_enabled"],
    "openai": ["openai", "enable_openai", "openai_enabled"],
    "railway": ["railway", "enable_railway", "railway_enabled"],
}

TRUTHY = {"1", "true", "yes", "y", "on", "enabled", "active"}
FALSY = {"0", "false", "no", "n", "off", "disabled", "inactive"}


def file_content_hash(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def _normalize_header(header: str) -> str:
    return re.sub(r"\s+", " ", (header or "").strip().lower())


def _build_header_map(headers: List[str]) -> Tuple[Dict[str, str], Dict[str, str]]:
    """Map canonical field names → actual column header in the file."""
    normalized = {_normalize_header(h): h for h in headers if h}
    mapping: Dict[str, str] = {}

    for field, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            if alias in normalized:
                mapping[field] = normalized[alias]
                break

    connector_map: Dict[str, str] = {}
    for connector, aliases in CONNECTOR_COLUMNS.items():
        for alias in aliases:
            if alias in normalized:
                connector_map[connector] = normalized[alias]
                break

    return mapping, connector_map


def _slugify(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "_", (value or "").lower()).strip("_")
    return slug[:64] or "imported_assignment"


def _parse_int(value: Any) -> Optional[int]:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).strip().replace(",", "").replace("$", "")
    if not text:
        return None
    try:
        return int(float(text))
    except (TypeError, ValueError):
        return None


def _parse_bool(value: Any) -> Optional[bool]:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in TRUTHY:
        return True
    if text in FALSY:
        return False
    return None


def _parse_status(value: Any) -> str:
    if value is None or value == "":
        return "active"
    text = str(value).strip().lower()
    if text in ("paused", "on hold", "on_hold", "hold"):
        return "paused"
    if text in ("archived", "done", "complete", "completed", "closed"):
        return "archived"
    if text in ("inactive", "disabled"):
        return "inactive"
    return "active"


def _row_to_assignment(
    row: Dict[str, Any],
    field_map: Dict[str, str],
    connector_map: Dict[str, str],
    row_number: int,
) -> Tuple[Optional[Dict[str, Any]], List[str], List[str]]:
    errors: List[str] = []
    warnings: List[str] = []

    def cell(field: str) -> Any:
        header = field_map.get(field)
        if not header:
            return None
        return row.get(header)

    name = cell("name")
    if name is None or str(name).strip() == "":
        # Allow id-only rows
        raw_id = cell("assignment_id")
        if raw_id is None or str(raw_id).strip() == "":
            errors.append(f"Row {row_number}: missing name and id")
            return None, errors, warnings
        name = str(raw_id).strip()

    assignment_id = cell("assignment_id")
    if assignment_id is None or str(assignment_id).strip() == "":
        assignment_id = _slugify(str(name))
        warnings.append(f"Row {row_number}: generated id '{assignment_id}' from name")
    else:
        assignment_id = _slugify(str(assignment_id).strip())

    metrics_config: Dict[str, Any] = {}
    for connector, header in connector_map.items():
        flag = _parse_bool(row.get(header))
        if flag:
            metrics_config[connector] = {"enabled": True}

    assignment = {
        "id": assignment_id,
        "assignment_id": assignment_id,
        "name": str(name).strip(),
        "description": str(cell("description") or "").strip(),
        "team_size": _parse_int(cell("team_size")),
        "monthly_burn_rate": _parse_int(cell("monthly_burn_rate")),
        "target_monthly_burn": _parse_int(cell("target_monthly_burn")),
        "status": _parse_status(cell("status")),
        "metrics_config": metrics_config,
        "_source_row": row_number,
    }

    if assignment["monthly_burn_rate"] is None and assignment["target_monthly_burn"] is None:
        warnings.append(f"Row {row_number} ({assignment_id}): no burn or target budget set")

    return assignment, errors, warnings


def _read_csv_rows(content: bytes) -> Tuple[List[Dict[str, Any]], List[str]]:
    errors: List[str] = []
    text = content.decode("utf-8-sig", errors="replace")
    # Try comma first, then semicolon (common in EU exports).
    for delimiter in (",", ";", "\t"):
        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
        if reader.fieldnames and len(reader.fieldnames) > 1:
            rows = [dict(r) for r in reader]
            if rows or reader.fieldnames:
                return rows, errors
    errors.append("Could not parse CSV — unrecognized delimiter or empty file")
    return [], errors


def _read_xlsx_rows(content: bytes) -> Tuple[List[Dict[str, Any]], List[str]]:
    errors: List[str] = []
    try:
        from openpyxl import load_workbook
    except ImportError:
        errors.append("Excel support requires openpyxl (pip install openpyxl)")
        return [], errors

    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        errors.append("Excel file is empty")
        return [], errors

    headers = [str(h).strip() if h is not None else "" for h in header_row]
    if not any(headers):
        errors.append("Excel file has no header row")
        return [], errors

    rows: List[Dict[str, Any]] = []
    for values in rows_iter:
        if not values or all(v is None or str(v).strip() == "" for v in values):
            continue
        row = {}
        for i, header in enumerate(headers):
            if not header:
                continue
            row[header] = values[i] if i < len(values) else None
        rows.append(row)
    return rows, errors


def parse_spreadsheet(
    content: bytes,
    filename: str,
) -> Dict[str, Any]:
    """
    Parse CSV or XLSX bytes into normalized assignment dicts.

    Returns a summary dict with assignments, errors, warnings, and metadata.
    Never raises for malformed rows — errors are collected per row.
    """
    ext = (filename or "").rsplit(".", 1)[-1].lower()
    file_hash = file_content_hash(content)

    result: Dict[str, Any] = {
        "valid": False,
        "file_hash": file_hash,
        "filename": filename,
        "format": ext,
        "rows_total": 0,
        "rows_parsed": 0,
        "rows_skipped": 0,
        "assignments": [],
        "errors": [],
        "warnings": [],
        "detected_columns": [],
    }

    if ext == "csv":
        raw_rows, read_errors = _read_csv_rows(content)
        result["errors"].extend(read_errors)
    elif ext in ("xlsx", "xlsm"):
        if ext == "xls":
            result["errors"].append(".xls (legacy Excel) is not supported — save as .xlsx")
            return result
        raw_rows, read_errors = _read_xlsx_rows(content)
        result["errors"].extend(read_errors)
    else:
        result["errors"].append(f"Unsupported file type: .{ext} (use .csv or .xlsx)")
        return result

    if result["errors"] and not raw_rows:
        return result

    result["rows_total"] = len(raw_rows)
    if not raw_rows:
        result["errors"].append("No data rows found in file")
        return result

    headers = list(raw_rows[0].keys())
    result["detected_columns"] = headers
    field_map, connector_map = _build_header_map(headers)

    if "name" not in field_map and "assignment_id" not in field_map:
        result["errors"].append(
            "Could not find a name or id column. "
            f"Detected headers: {', '.join(headers[:12])}"
        )
        return result

    assignments: List[Dict[str, Any]] = []
    seen_ids: Dict[str, int] = {}

    for idx, row in enumerate(raw_rows, start=2):  # row 1 = header
        assignment, row_errors, row_warnings = _row_to_assignment(
            row, field_map, connector_map, idx
        )
        result["errors"].extend(row_errors)
        result["warnings"].extend(row_warnings)

        if assignment is None:
            result["rows_skipped"] += 1
            continue

        aid = assignment["assignment_id"]
        if aid in seen_ids:
            suffix = seen_ids[aid] + 1
            seen_ids[aid] = suffix
            new_id = f"{aid}_{suffix}"
            result["warnings"].append(
                f"Row {idx}: duplicate id '{aid}' renamed to '{new_id}'"
            )
            assignment["assignment_id"] = new_id
            assignment["id"] = new_id
        else:
            seen_ids[aid] = 1

        assignments.append(assignment)
        result["rows_parsed"] += 1

    result["assignments"] = assignments
    result["valid"] = len(assignments) > 0 and len(result["errors"]) == 0
    # Partial success: some rows OK even with row-level errors elsewhere
    if assignments and result["errors"]:
        result["valid"] = True
        result["warnings"].append(
            f"Imported {len(assignments)} rows with {len(result['errors'])} row errors"
        )

    logger.info(
        "Spreadsheet parsed: %s rows=%d parsed=%d errors=%d",
        filename,
        result["rows_total"],
        result["rows_parsed"],
        len(result["errors"]),
    )
    return result
